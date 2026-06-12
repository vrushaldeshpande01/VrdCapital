"""
Excel (XLSX) report generator using openpyxl.
Produces formatted spreadsheets with VrdCapital branding.
"""
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

NAVY_HEX  = "1a237e"
TEAL_HEX  = "00695c"
LIGHT_HEX = "e8eaf6"
WHITE_HEX = "FFFFFF"
GREEN_HEX = "2e7d32"
RED_HEX   = "c62828"

def _navy_fill():  return PatternFill("solid", fgColor=NAVY_HEX)
def _light_fill(): return PatternFill("solid", fgColor=LIGHT_HEX)
def _white_fill(): return PatternFill("solid", fgColor=WHITE_HEX)

def _header_font(): return Font(name="Calibri", bold=True, color=WHITE_HEX, size=10)
def _bold_font():   return Font(name="Calibri", bold=True, size=10)
def _body_font():   return Font(name="Calibri", size=10)
def _title_font():  return Font(name="Calibri", bold=True, size=14, color=NAVY_HEX)

def _thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _set_header_row(ws, row, headers, col_widths=None):
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.fill = _navy_fill()
        cell.font = _header_font()
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _thin_border()
        if col_widths and i <= len(col_widths):
            ws.column_dimensions[get_column_letter(i)].width = col_widths[i-1]

def _set_data_row(ws, row, values, even=True):
    fill = _light_fill() if even else _white_fill()
    for i, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=i, value=v)
        cell.fill = fill
        cell.font = _body_font()
        cell.border = _thin_border()
        cell.alignment = Alignment(vertical="center")

def _title_section(ws, title, subtitle):
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 20
    ws["A1"] = "VrdCapital — " + title
    ws["A1"].font = _title_font()
    ws["A2"] = subtitle
    ws["A2"].font = Font(name="Calibri", size=10, color="757575")
    ws["A3"] = f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}"
    ws["A3"].font = Font(name="Calibri", size=9, color="9e9e9e", italic=True)

def _fmt_inr(val, prefix="₹") -> str:
    try:
        n = float(val or 0)
        if abs(n) >= 1e7: return f"{prefix}{n/1e7:.2f} Cr"
        if abs(n) >= 1e5: return f"{prefix}{n/1e5:.2f} L"
        return f"{prefix}{n:,.2f}"
    except Exception: return "—"

def _pct(val) -> str:
    try:
        n = float(val or 0)
        return f"{'+' if n>=0 else ''}{n:.2f}%"
    except Exception: return "—"


def generate_portfolio_summary_xlsx(data: dict, period_label: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Portfolio Summary"
    _title_section(ws, "Portfolio Summary", f"Period: {period_label}")

    summary = data.get("summary", {})
    holdings = data.get("holdings", [])
    history  = data.get("history", [])

    # KPI section
    row = 5
    ws.cell(row=row, column=1, value="Key Metrics").font = Font(name="Calibri", bold=True, size=11, color=NAVY_HEX)
    row += 1
    kpis = [
        ("Total Value", _fmt_inr(summary.get("total_value"))),
        ("Total Invested", _fmt_inr(summary.get("total_invested"))),
        ("Total P&L", _fmt_inr(summary.get("total_pnl"))),
        ("Return %", _pct(summary.get("total_pnl_pct"))),
        ("Day P&L", _fmt_inr(summary.get("day_pnl"))),
        ("Holdings Count", str(summary.get("num_holdings","—"))),
    ]
    for label, value in kpis:
        ws.cell(row=row, column=1, value=label).font = Font(name="Calibri", bold=True, size=10)
        ws.cell(row=row, column=2, value=value).font = Font(name="Calibri", size=10)
        row += 1

    row += 1
    # Holdings table
    if holdings:
        ws.cell(row=row, column=1, value="Holdings").font = Font(name="Calibri", bold=True, size=11, color=NAVY_HEX)
        row += 1
        headers = ["Symbol","Exchange","Quantity","Avg Cost","LTP","P&L","P&L %","Weight %"]
        widths  = [12, 10, 10, 14, 14, 14, 10, 10]
        _set_header_row(ws, row, headers, widths)
        row += 1
        for i, h in enumerate(holdings):
            _set_data_row(ws, row, [
                h.get("symbol",""), h.get("exchange","NSE"), h.get("quantity",0),
                _fmt_inr(h.get("average_cost")), _fmt_inr(h.get("last_price")),
                _fmt_inr(h.get("pnl")), _pct(h.get("pnl_pct")),
                f"{float(h.get('weight_pct',0)):.1f}%",
            ], even=i % 2 == 0)
            row += 1
        row += 1

    # History table (separate sheet)
    if history:
        ws2 = wb.create_sheet("Value History")
        _title_section(ws2, "Portfolio Value History", f"Period: {period_label}")
        r = 5
        _set_header_row(ws2, r, ["Date","Portfolio Value","Day Change"], [18, 18, 18])
        r += 1
        rows_sorted = sorted(history, key=lambda x: x.get("date",""), reverse=True)
        for i, row_data in enumerate(rows_sorted):
            prev = float(rows_sorted[i+1]["portfolio_value"]) if i+1 < len(rows_sorted) else float(row_data["portfolio_value"])
            chg = float(row_data["portfolio_value"]) - prev
            _set_data_row(ws2, r, [
                row_data.get("date",""), _fmt_inr(row_data.get("portfolio_value")),
                _fmt_inr(chg) if i < len(rows_sorted)-1 else "—"
            ], even=i % 2 == 0)
            r += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_order_history_xlsx(data: dict, period_label: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Order History"
    _title_section(ws, "Order History", f"Period: {period_label}")

    orders = data.get("orders", [])
    row = 5
    headers = ["Date","Symbol","Exchange","Side","Type","Quantity","Price","Broker","Status"]
    widths  = [18, 12, 10, 8, 10, 10, 14, 12, 14]
    _set_header_row(ws, row, headers, widths)
    row += 1
    for i, o in enumerate(orders):
        dt = (o.get("placed_at","")[:16] or "").replace("T"," ")
        _set_data_row(ws, row, [
            dt, o.get("symbol",""), o.get("exchange",""), o.get("side",""),
            o.get("price_type",""), o.get("quantity",0),
            _fmt_inr(o.get("average_price") or o.get("price")),
            o.get("broker",""), o.get("status",""),
        ], even=i % 2 == 0)
        row += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_csv(data: dict, report_type: str) -> bytes:
    """Generic CSV fallback — exports holdings or orders as CSV."""
    import csv
    buf = io.StringIO()
    writer = csv.writer(buf)

    if report_type == "order_history":
        orders = data.get("orders", [])
        if orders:
            writer.writerow(orders[0].keys())
            for o in orders:
                writer.writerow(o.values())
    else:
        holdings = data.get("holdings", [])
        if holdings:
            writer.writerow(holdings[0].keys())
            for h in holdings:
                writer.writerow(h.values())

    return buf.getvalue().encode("utf-8")
